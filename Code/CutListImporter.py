import os
import re
from Utils import frac_to_value
import openpyxl

def import_text_list(inputFile):
    cutsDict = {}
    currentCategory = []

    with open(inputFile) as file:
        for line_num, l in enumerate(file, start=1):
            line = l.strip()
            if len(line) == 0:
                continue
            try:
                if line[-1] == ":":
                    category = re.sub(r"[^0-9x\s]", "", line)
                    category = (int(category[0 : category.index("x")]), int(category[category.index("x") + 1 :]))
                    if category not in cutsDict:
                        cutsDict[category] = []
                    currentCategory = cutsDict[category]   
                else:
                    count = line[0 : line.index("x")]
                    count = int(re.sub(r'[^0-9]', '', count))

                    note = ""
                    if "#" in line:
                        note = line[line.index("#") + 1:].strip()
                        measurement = frac_to_value(line[line.index("x") + 1:line.index("#")])
                    else:
                        measurement = frac_to_value(line[line.index("x") + 1:])
                    for _ in range(count):
                        currentCategory.append((measurement, note))
            except Exception as e:
                raise Exception(f"Error on line {line_num}: \"{line}\", {e}")
            
    numCategories = len(cutsDict)
    numCuts = 0
    for c in cutsDict.values():
        numCuts += len(c) - 2
    for key in cutsDict:
        cutsDict[key] = sorted(cutsDict[key], reverse=True)
    
    print(f"\nImported cut list \"{os.path.basename(inputFile)}\"\n{numCategories} categories, {numCuts} cuts\n")
    
    sorted_categories = dict(sorted(cutsDict.items(), key=lambda category: (category[0], category[1])))
    return sorted_categories

def import_spreadsheet(input_file):
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
    cuts_dict = {}
    current_category = []

    for i in range(1, sheet.max_column + 1):
        category_value = re.sub(r"[^0-9x\s]", "", str(sheet.cell(1, i).value))
        
        
        if not "x" in category_value:
            continue
        category_value = (int(category_value[0:category_value.index("x")]), int(category_value[category_value.index("x") + 1:]))
        if not category_value in cuts_dict:
            cuts_dict[category_value] = []

        current_category = cuts_dict[category_value]
        for j in range(2, sheet.max_row + 1):
            try:
                measurement = float(sheet.cell(j, i).value)
            except Exception:
                continue
    
            count = sheet.cell(j, i + 1).value
            if count == "":
                count = 1
            else:
                try:
                    count = int(count)
                except Exception:
                    continue
                       
            note = sheet.cell(j, i + 2).value
            if note == None:
                note = ""
            for _ in range(count):
                current_category.append((measurement, note))

        if len(current_category) == 0:
            cuts_dict.popitem()

    numCategories = len(cuts_dict)
    numCuts = 0
    for c in cuts_dict.values():
        numCuts += len(c) - 2
    for key in cuts_dict:
        cuts_dict[key] = sorted(cuts_dict[key], reverse=True)
    
    print(f"\nImported cut list \"{os.path.basename(input_file)}\"\n{numCategories} categories, {numCuts} cuts\n")
    
    sorted_categories = dict(sorted(cuts_dict.items(), key=lambda category: (category[0], category[1])))

    return sorted_categories